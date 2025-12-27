"""Pivot table export helpers."""

from __future__ import annotations

from copy import copy
from io import BytesIO
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analysis.views.pivot_utils import NestedPivotData, format_key_label


def build_nested_pivot_export_df(data: NestedPivotData) -> pd.DataFrame:
    """Flatten nested pivot data into a tabular export DataFrame."""
    col_defs: List[Dict[str, str]] = []
    for col_key, col_tuple in zip(data.col_keys, data.col_key_tuples):
        col_label = format_key_label(col_key)
        if not col_label:
            col_label = "总体"
        for value_col in data.value_cols:
            if len(data.value_cols) > 1:
                label = f"{col_label} | {value_col}"
            else:
                label = col_label or value_col
            col_defs.append(
                {
                    "label": label,
                    "col_tuple": col_tuple,
                    "value_col": value_col,
                }
            )

    rows: List[Dict[str, object]] = []
    for row_key, row_tuple in zip(data.row_keys, data.row_key_tuples):
        base = {col: row_key.get(col, "") for col in data.row_key_cols}
        for agg_name in data.agg_names:
            row: Dict[str, object] = dict(base)
            row["统计量"] = agg_name
            for col_def in col_defs:
                row[col_def["label"]] = data.values.get(
                    (row_tuple, col_def["col_tuple"], col_def["value_col"], agg_name)
                )
            rows.append(row)

    return pd.DataFrame(rows)


def nested_pivot_to_excel_bytes(data: NestedPivotData) -> bytes:
    """Render nested pivot data into an Excel workbook and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot"

    row_cols = list(data.row_key_cols or [])
    col_cols = list(data.col_key_cols or [])
    value_cols = list(data.value_cols or [])
    agg_names = list(data.agg_names or [])
    row_keys = list(data.row_keys or [])
    row_key_tuples = list(data.row_key_tuples or [])
    col_keys = list(data.col_keys or [])
    col_key_tuples = list(data.col_key_tuples or [])

    if not value_cols:
        value_cols = ["值"]
    if not agg_names:
        agg_names = ["-"]

    header_rows = 2
    row_header_cols = len(row_cols) + 1  # + 统计量
    data_start_col = row_header_cols + 1
    data_start_row = header_rows + 1
    group_size = max(len(agg_names), 1)

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    subheader_fill = PatternFill("solid", fgColor="E6E6E6")
    row_header_fill = PatternFill("solid", fgColor="F2F2F2")
    agg_fill = PatternFill("solid", fgColor="F7F7F7")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    def apply_style(
        cell,
        fill=None,
        font=None,
        alignment=None,
        number_format=None,
    ) -> None:
        """Apply common style attributes to a worksheet cell."""
        cell.border = border
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if number_format is not None:
            cell.number_format = number_format

    def update_width(widths: dict[int, int], col: int, value: object) -> None:
        """Update the max width tracker for an Excel column."""
        if value is None:
            return
        text = str(value)
        if "\n" in text:
            text = max(text.split("\n"), key=len)
        widths[col] = max(widths.get(col, 0), len(text))

    col_groups: List[Dict[str, object]] = []
    for col_key, col_tuple in zip(col_keys, col_key_tuples):
        col_label = format_key_label(col_key)
        if not col_label:
            col_label = "总体"
        col_groups.append(
            {"col_tuple": col_tuple, "label": col_label}
        )
    if not col_groups:
        col_groups = [{"col_tuple": (), "label": "总体"}]

    max_widths: dict[int, int] = {}

    # Header row: row dimensions + 统计量
    for idx, col_name in enumerate(row_cols, start=1):
        cell = ws.cell(row=1, column=idx, value=col_name)
        apply_style(cell, header_fill, header_font, center_align)
        update_width(max_widths, idx, col_name)
        ws.merge_cells(start_row=1, end_row=2, start_column=idx, end_column=idx)

    stat_col = len(row_cols) + 1
    stat_cell = ws.cell(row=1, column=stat_col, value="统计量")
    apply_style(stat_cell, header_fill, header_font, center_align)
    update_width(max_widths, stat_col, "统计量")
    ws.merge_cells(
        start_row=1, end_row=2, start_column=stat_col, end_column=stat_col
    )

    # Column headers (two-row)
    col_idx = data_start_col
    for group in col_groups:
        start_col = col_idx
        for value_col in value_cols:
            sub_cell = ws.cell(row=2, column=col_idx, value=value_col)
            apply_style(sub_cell, subheader_fill, header_font, center_align)
            update_width(max_widths, col_idx, value_col)
            col_idx += 1
        end_col = col_idx - 1
        if start_col < end_col:
            ws.merge_cells(
                start_row=1, end_row=1, start_column=start_col, end_column=end_col
            )
        group_cell = ws.cell(row=1, column=start_col, value=group["label"])
        apply_style(group_cell, header_fill, header_font, center_align)
        update_width(max_widths, start_col, group["label"])

    # Data rows
    for row_idx, row_key in enumerate(row_keys):
        row_tuple = row_key_tuples[row_idx] if row_idx < len(row_key_tuples) else ()
        for agg_idx, agg_name in enumerate(agg_names):
            excel_row = data_start_row + row_idx * group_size + agg_idx
            if agg_idx == 0:
                for col_idx, col_name in enumerate(row_cols, start=1):
                    val = row_key.get(col_name, "")
                    cell = ws.cell(row=excel_row, column=col_idx, value=val)
                    apply_style(cell, row_header_fill, None, left_align)
                    update_width(max_widths, col_idx, val)

            agg_cell = ws.cell(
                row=excel_row, column=stat_col, value=agg_name
            )
            apply_style(agg_cell, agg_fill, header_font, left_align)
            update_width(max_widths, stat_col, agg_name)

            data_col = data_start_col
            for group in col_groups:
                col_tuple = group["col_tuple"]
                for value_col in value_cols:
                    val = data.values.get(
                        (row_tuple, col_tuple, value_col, agg_name)
                    )
                    if val is None or pd.isna(val):
                        cell_val = "-"
                        cell = ws.cell(row=excel_row, column=data_col, value=cell_val)
                        apply_style(cell, None, None, center_align)
                        update_width(max_widths, data_col, cell_val)
                    else:
                        try:
                            num_val = round(float(val), 4)
                            cell = ws.cell(
                                row=excel_row, column=data_col, value=num_val
                            )
                            apply_style(
                                cell,
                                None,
                                None,
                                right_align,
                                number_format="0.000",
                            )
                            update_width(
                                max_widths, data_col, f"{num_val:.3f}"
                            )
                        except Exception:
                            cell_val = str(val)
                            cell = ws.cell(
                                row=excel_row, column=data_col, value=cell_val
                            )
                            apply_style(cell, None, None, left_align)
                            update_width(max_widths, data_col, cell_val)
                    data_col += 1

    # Merge row headers (hierarchical)
    if row_cols and row_keys:
        for level, col_name in enumerate(row_cols):
            block_start = 0
            prev_key = row_keys[0]
            for idx in range(1, len(row_keys) + 1):
                if idx == len(row_keys):
                    next_key = None
                else:
                    next_key = row_keys[idx]

                def prefix_changed() -> bool:
                    """Return True when the row-key prefix changes at this level."""
                    if next_key is None:
                        return True
                    for j in range(level):
                        if prev_key.get(row_cols[j]) != next_key.get(row_cols[j]):
                            return True
                    return prev_key.get(col_name) != next_key.get(col_name)

                if prefix_changed():
                    start_row = data_start_row + block_start * group_size
                    end_row = data_start_row + idx * group_size - 1
                    if start_row < end_row:
                        ws.merge_cells(
                            start_row=start_row,
                            end_row=end_row,
                            start_column=level + 1,
                            end_column=level + 1,
                        )
                    block_start = idx
                    if next_key is not None:
                        prev_key = next_key

    # Borders for all cells in range
    last_row = (
        data_start_row + len(row_keys) * group_size - 1
        if row_keys
        else header_rows
    )
    last_col = data_start_col + len(col_groups) * len(value_cols) - 1
    for r in range(1, last_row + 1):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            if cell.border is None or cell.border == Border():
                cell.border = border
            alignment = cell.alignment
            if alignment is None:
                cell.alignment = Alignment(vertical="center")
            else:
                updated_alignment = copy(alignment)
                updated_alignment.vertical = "center"
                cell.alignment = updated_alignment

    # Freeze panes
    ws.freeze_panes = ws.cell(
        row=header_rows + 1, column=row_header_cols + 1
    )

    # Column widths
    for col_idx in range(1, last_col + 1):
        col_letter = get_column_letter(col_idx)
        width = max_widths.get(col_idx, 8) + 2
        ws.column_dimensions[col_letter].width = min(max(width, 8), 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
