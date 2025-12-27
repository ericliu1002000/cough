"""Export helpers for subject profile data."""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from typing import Dict

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def to_excel_bytes(tables: Dict[str, pd.DataFrame]) -> bytes:
    """Return a multi-sheet Excel export for subject profile tables."""
    output = BytesIO()
    sheet_names: Dict[str, int] = {}
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for table_name, df in tables.items():
            safe_name = "".join(
                ch for ch in str(table_name) if ch not in r"[]:*?/\\"
            ).strip()
            safe_name = safe_name[:31] if safe_name else "Sheet"
            if safe_name in sheet_names:
                sheet_names[safe_name] += 1
                suffix = sheet_names[safe_name]
                base = safe_name[: max(0, 31 - len(str(suffix)) - 1)]
                safe_name = f"{base}_{suffix}"
            else:
                sheet_names[safe_name] = 1
            df.to_excel(writer, sheet_name=safe_name, index=False)
    output.seek(0)
    return output.getvalue()


def to_excel_sections_bytes(
    tables: Dict[str, pd.DataFrame], subject_id: str | None = None
) -> bytes:
    """Return a single-sheet Excel export with sectioned tables."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Subject"

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    summary_rows = [
        ("受试者ID", subject_id or ""),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("表数量", str(len(tables))),
        ("总行数", str(sum(len(df) for df in tables.values()))),
    ]
    row = 1
    for label, value in summary_rows:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1

    for table_name, df in tables.items():
        columns = list(df.columns)
        data_rows = max(len(df), 1)
        start_row = row
        header_row = row
        data_start = header_row + 1
        end_row = header_row + data_rows
        last_col = 1 + max(len(columns), 1)

        ws.merge_cells(
            start_row=start_row, start_column=1, end_row=end_row, end_column=1
        )
        table_cell = ws.cell(row=start_row, column=1, value=str(table_name))
        table_cell.font = header_font
        table_cell.fill = header_fill
        table_cell.alignment = center_align

        for col_idx, col_name in enumerate(columns, start=2):
            cell = ws.cell(row=header_row, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        if df.empty:
            msg_cell = ws.cell(row=data_start, column=2, value="(no rows)")
            msg_cell.alignment = center_align
            if last_col > 2:
                ws.merge_cells(
                    start_row=data_start,
                    start_column=2,
                    end_row=data_start,
                    end_column=last_col,
                )
        else:
            for r_idx, row_data in enumerate(df.itertuples(index=False), start=data_start):
                for c_idx, value in enumerate(row_data, start=2):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    col_name = columns[c_idx - 2]
                    if pd.api.types.is_numeric_dtype(df[col_name]):
                        cell.alignment = right_align
                    else:
                        cell.alignment = left_align

        for r in range(start_row, end_row + 1):
            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue
                cell.border = border
                if c == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                elif r == header_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

        row = end_row + 2

    ws.column_dimensions["A"].width = 18
    for col_idx in range(2, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def to_csv_sections_bytes(tables: Dict[str, pd.DataFrame]) -> bytes:
    """Return a CSV export with table sections separated by headers."""
    output = StringIO()
    writer = csv.writer(output)
    first = True
    for table_name, df in tables.items():
        if not first:
            writer.writerow([])
        writer.writerow([f"TABLE: {table_name}"])
        if df.empty:
            writer.writerow(["(no rows)"])
            first = False
            continue
        writer.writerow(list(df.columns))
        for row in df.itertuples(index=False, name=None):
            writer.writerow(list(row))
        first = False
    return output.getvalue().encode("utf-8-sig")
