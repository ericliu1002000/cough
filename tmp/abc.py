import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# 1. 定义表格数据
data = {
    "序号": list(range(1, 13)) + ["分拣总数", "每斤采购价格", "上货斤数", "上货金额", "泥重", "杂质率", "每斤毛利润", "每斤净利润", "总利润（毛）", "总利润（净）"],
    "规格": ["豆", "1+", "1.5+", "2+", "2.5+", "3+", "4+", "5+", "6+", "7+", "8+", "9+"] + ["", "", "", "", "", "", "", "", "", ""],
    "重量": [""] * 12 + ["", "", "", "", "", "", "", "", "", ""],
    "单价": [""] * 12 + ["", "", "", "", "", "", "", "", "", ""],
    "金额": [""] * 12 + ["", "", "", "", "", "", "", "", "", ""],
    "备注": ["-"] * 12 + ["", "", "", "", "", "", "", "", "", ""],
    "比例": [""] * 12 + ["", "", "", "", "", "", "", "", "", ""]
}

df = pd.DataFrame(data)

# 2. 保存到 Excel
file_path = "采购利润表.xlsx"
df.to_excel(file_path, index=False, engine="openpyxl")

# 3. 用 openpyxl 打开并设置公式和样式
wb = load_workbook(file_path)
ws = wb.active

# 样式定义
style_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 浅绿色
style_yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # 浅黄色
style_blue = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")    # 浅蓝色

# 固定列（规格列 B）
for row in range(2, ws.max_row + 1):
    ws[f"B{row}"].fill = style_green

# 手动填写列（重量 C, 单价 D, 采购价 B16, 上货斤数 D16, 泥重 B17）
manual_cols = ["C", "D"]
for col in manual_cols:
    for row in range(3, 15):  # 数据行
        ws[f"{col}{row}"].fill = style_yellow
ws["B16"].fill = style_yellow  # 每斤采购价格
ws["D16"].fill = style_yellow  # 上货斤数
ws["B17"].fill = style_yellow  # 泥重

# 自动计算列（金额 E, 比例 G, 汇总行）
auto_cols = ["E", "G"]
for col in auto_cols:
    for row in range(3, 15):
        ws[f"{col}{row}"].fill = style_blue

# 4. 设置公式
# 金额 = 重量 * 单价
for row in range(3, 15):
    ws[f"E{row}"] = f"=C{row}*D{row}"

# 比例 = 金额 / 销售总额
for row in range(3, 15):
    ws[f"G{row}"] = f"=E{row}/E$15"

# 分拣总数 = SUM(重量)
ws["C15"] = "=SUM(C3:C14)"

# 销售总额 = SUM(金额)
ws["E15"] = "=SUM(E3:E14)"

# 3+以上比例 = SUM(G7:G10)
ws["G15"] = "=SUM(G7:G10)"

# 上货金额 = 每斤采购价格 * 上货斤数
ws["F16"] = "=B16*D16"

# 杂质率 = 泥重 / 上货斤数
ws["D17"] = "=B17/D16"

# 每斤毛利润 = (销售总额 - 上货金额) / 分拣总数
ws["B18"] = "=(E15-F16)/C15"

# 总利润（毛）= 销售总额 - 上货金额
ws["F18"] = "=E15-F16"

# 每斤净利润（示例公式，可根据实际成本调整）
ws["D18"] = "=F18/C15"

# 总利润（净）= 每斤净利润 * 分拣总数
ws["F18"] = "=D18*C15"

# 设置百分比格式
for row in range(3, 16):
    ws[f"G{row}"].number_format = "0.00%"
ws["D17"].number_format = "0.00%"

# 5. 保存文件
wb.save(file_path)

print(f"Excel 模板已生成：{file_path}")